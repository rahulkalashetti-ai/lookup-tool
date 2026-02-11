"""Export scan results to Excel and PDF (PRD §3.2)."""
from pathlib import Path
import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from ai_scanner import STATUS_AVAILABLE, STATUS_UNAVAILABLE, STATUS_REVIEW

def export_excel(result_df: pd.DataFrame, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    result_df.to_excel(out_path, index=False, engine="openpyxl")
    wb = load_workbook(out_path)
    ws = wb.active
    # Header style
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    # Status column coloring (Availability is last-1 or last)
    avail_col = None
    for i, c in enumerate(ws[1], 1):
        if c.value == "Availability":
            avail_col = i
            break
    if avail_col:
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=avail_col).value
            if val == STATUS_AVAILABLE:
                ws.cell(row=row, column=avail_col).fill = green
            elif val == STATUS_UNAVAILABLE:
                ws.cell(row=row, column=avail_col).fill = red
            else:
                ws.cell(row=row, column=avail_col).fill = yellow
    wb.save(out_path)

def export_pdf(result_df: pd.DataFrame, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(str(out_path), pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph("Tool Availability Scan Results", styles["Title"]))
    elements.append(Spacer(1, 0.2*inch))
    # Convert to strings safely
    df_str = result_df.copy()
    for c in df_str.columns:
        df_str[c] = df_str[c].astype(object).fillna("").astype(str).str[:50]
    data = [list(df_str.columns)] + df_str.values.tolist()
    chunk = 25
    for i in range(0, len(data), chunk):
        t = Table(data[i:i+chunk])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(t)
        if i + chunk < len(data):
            elements.append(Spacer(1, 0.3*inch))
    doc.build(elements)
